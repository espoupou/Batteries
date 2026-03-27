from __future__ import annotations

import re
import math
from copy import copy
from pathlib import Path
from openpyxl import Workbook, load_workbook
from typing import Optional

import numpy as np
import pandas as pd


# =========================
# Global eis core variables
# =========================

EXCEL_FILE_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}


# =========================
# Helpers
# =========================


def preprocess_eclab_text_file(
    input_path: str | Path,
    output_path: str | Path | None = None,
    *,
    skip_first_line: bool = True,
    replace_comma: bool = True,
    keep_n_cols: int = 3,
    negate_col3: bool = False,
    stop_first_col: float | None = None,
    encoding: str = "utf-8",
) -> Path:
    """
    Lit un fichier texte, garde les N premières colonnes, applique éventuellement :
      - remplacement virgule -> point
      - inversion signe colonne 3
      - arrêt quand la 1re colonne atteint une valeur
    et écrit le résultat dans un nouveau fichier.
    """
    input_path = Path(input_path)

    if output_path is None:
        output_path = input_path.with_name(f"{input_path.stem}_OUTPUT{input_path.suffix}")
    else:
        output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with input_path.open("r", encoding=encoding, errors="ignore") as fin, \
         output_path.open("w", encoding=encoding) as fout:

        if skip_first_line:
            next(fin, None)

        start_line = 2 if skip_first_line else 1

        for line_num, line in enumerate(fin, start=start_line):
            line = line.strip()
            if not line:
                continue

            if replace_comma:
                line = line.replace(",", ".")

            parts = line.split()

            if len(parts) < keep_n_cols:
                raise ValueError(
                    f"{input_path.name} | ligne {line_num} n'a pas {keep_n_cols} colonnes: {line!r}"
                )

            parts = parts[:keep_n_cols]

            if stop_first_col is not None:
                try:
                    first_val = float(parts[0])
                except ValueError as e:
                    raise ValueError(
                        f"{input_path.name} | ligne {line_num} a une 1re colonne non numérique: {parts[0]!r}"
                    ) from e

                if math.isclose(first_val, stop_first_col, rel_tol=0.0, abs_tol=1e-12) or (first_val <= stop_first_col):
                    break

            if negate_col3:
                try:
                    parts[2] = str(-float(parts[2]))
                except ValueError as e:
                    raise ValueError(
                        f"{input_path.name} | ligne {line_num} a une 3e colonne non numérique: {parts[2]!r}"
                    ) from e

            fout.write("\t".join(parts) + "\n")

    return output_path


def preprocess_eclab_text_to_dataframe(
    input_path: str | Path,
    *,
    skip_first_line: bool = True,
    replace_comma: bool = True,
    keep_n_cols: int = 3,
    negate_col3: bool = False,
    stop_first_col: float | None = None,
    encoding: str = "utf-8",
) -> pd.DataFrame:
    """
    Prétraite un fichier texte EIS puis le relit comme DataFrame normalisé
    avec colonnes: freq, R, Im
    """
    input_path = Path(input_path)
    temp_path = input_path.with_name(f"{input_path.stem}__preview_tmp{input_path.suffix}")

    out_path = preprocess_eclab_text_file(
        input_path=input_path,
        output_path=temp_path,
        skip_first_line=skip_first_line,
        replace_comma=replace_comma,
        keep_n_cols=keep_n_cols,
        negate_col3=negate_col3,
        stop_first_col=stop_first_col,
        encoding=encoding,
    )

    try:
        df = pd.read_csv(
            out_path,
            sep=r"\s+",
            header=None,
            names=["freq", "R", "Im"],
            engine="python",
        )

        for c in ["freq", "R", "Im"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")

        df = df.dropna(subset=["freq", "R", "Im"]).reset_index(drop=True)
        return df

    finally:
        try:
            out_path.unlink(missing_ok=True)
        except Exception:
            pass


def _truncate_normalized_df_at_stop(df: pd.DataFrame, stop_first_col: float | None) -> pd.DataFrame:
    """
    Tronque un DataFrame normalisé (freq, R, Im) quand la première colonne
    atteint ou franchit la valeur demandée, en conservant la ligne seuil.
    """
    if stop_first_col is None or df.empty:
        return df.reset_index(drop=True).copy()

    freq_values = pd.to_numeric(df["freq"], errors="coerce")
    if freq_values.isna().all():
        return df.reset_index(drop=True).copy()

    keep_until = len(df)
    prev_val = None
    scan_direction = 0  # +1 ascending, -1 descending

    for idx, current in enumerate(freq_values.to_numpy(dtype=float)):
        if prev_val is not None and scan_direction == 0 and not math.isclose(current, prev_val, rel_tol=0.0, abs_tol=1e-18):
            scan_direction = +1 if current > prev_val else -1

        reached = math.isclose(current, stop_first_col, rel_tol=0.0, abs_tol=1e-12)
        if scan_direction == -1:
            reached = reached or (current <= stop_first_col)
        elif scan_direction == +1:
            reached = reached or (current >= stop_first_col)

        if reached:
            keep_until = idx + 1
            break

        prev_val = current

    return df.iloc[:keep_until].copy().reset_index(drop=True)


def is_excel_file(path: str | Path) -> bool:
    return Path(path).suffix.lower() in EXCEL_FILE_EXTENSIONS


def preprocess_eis_dataframe(
    table_df: pd.DataFrame,
    *,
    keep_n_cols: int = 3,
    negate_col3: bool = False,
    stop_first_col: float | None = None,
) -> pd.DataFrame:
    """
    Normalise une table EIS (par exemple une feuille Excel) vers les colonnes:
        freq, R, Im

    Si des headers compatibles sont présents, ils sont détectés.
    Sinon, on utilise les 3 premières colonnes numériques exploitables.
    """
    if table_df is None:
        raise ValueError("Aucune table fournie.")

    work = table_df.copy()
    work = work.dropna(axis=0, how="all")
    work = work.dropna(axis=1, how="all")

    if work.empty:
        raise ValueError("La table est vide.")

    selected = {}
    im_sign = +1

    for col in work.columns:
        role, sign = _classify_header(col)
        if role and role not in selected:
            selected[role] = col
            if role == "Im":
                im_sign = sign if sign is not None else +1

    if {"freq", "R", "Im"}.issubset(selected):
        out = work[[selected["freq"], selected["R"], selected["Im"]]].copy()
        out.columns = ["freq", "R", "Im"]
    else:
        n_cols = max(int(keep_n_cols), 3)
        if work.shape[1] < 3:
            raise ValueError("La feuille ne contient pas au moins 3 colonnes exploitables.")
        out = work.iloc[:, :n_cols].copy().iloc[:, :3]
        out.columns = ["freq", "R", "Im"]

    for c in ["freq", "R", "Im"]:
        out[c] = out[c].astype(str).str.replace(",", ".", regex=False)
        out[c] = pd.to_numeric(out[c], errors="coerce")

    out = out.dropna(subset=["freq", "R", "Im"]).reset_index(drop=True)

    if {"freq", "R", "Im"}.issubset(selected) and im_sign == -1:
        out["Im"] = -out["Im"]

    if negate_col3:
        out["Im"] = -out["Im"]

    out = _truncate_normalized_df_at_stop(out, stop_first_col)
    return out.reset_index(drop=True)


def read_eis_excel_workbook(input_path: str | Path) -> dict[str, pd.DataFrame]:
    """
    Charge toutes les feuilles d'un classeur Excel EIS dans l'ordre du workbook.
    Les tables sont gardées telles quelles pour permettre la navigation feuille par feuille.
    """
    input_path = Path(input_path)
    if not is_excel_file(input_path):
        raise ValueError(f"Le fichier n'est pas un classeur Excel supporté: {input_path}")

    engine = "openpyxl" if input_path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"} else None
    sheets = pd.read_excel(input_path, sheet_name=None, dtype=str, engine=engine)

    cleaned = {}
    for sheet_name, df in sheets.items():
        if df is None:
            continue
        block = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        if block.empty:
            continue
        cleaned[str(sheet_name)] = block.reset_index(drop=True)

    if not cleaned:
        raise ValueError("Aucune feuille exploitable trouvée dans le classeur Excel.")

    return cleaned


def export_processed_eis_workbook(
    processed_sheets: dict[str, pd.DataFrame],
    out_path: str | Path,
) -> Path:
    """
    Exporte plusieurs jeux EIS traités dans un seul classeur Excel.
    Les noms de feuilles d'origine sont conservés autant que possible.
    """
    if not processed_sheets:
        raise ValueError("Aucune feuille traitée à exporter.")

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    used_names = set()
    summary_rows = []

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for original_name, df in processed_sheets.items():
            sheet_name = make_unique_sheet_name(str(original_name), used_names)
            used_names.add(sheet_name)

            out_df = df[["freq", "R", "Im"]].copy()
            out_df.columns = ["freq_Hz", "Zprime_Ohm", "Zdoubleprime_Ohm"]
            out_df.to_excel(writer, sheet_name=sheet_name, index=False)

            summary_row = {
                "source_sheet": str(original_name),
                "export_sheet": sheet_name,
                "n_points": int(len(df)),
            }
            if not df.empty:
                summary_row["f_start_Hz"] = float(df["freq"].iloc[0])
                summary_row["f_end_Hz"] = float(df["freq"].iloc[-1])
            summary_rows.append(summary_row)

        summary_name = make_unique_sheet_name("Summary", used_names)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name=summary_name, index=False)

    return out_path


# =========================
# Lecture EIS
# =========================

def _clean_header(s: str) -> str:
    s = str(s).strip().lower()
    s = s.replace("−", "-").replace("–", "-")
    s = s.replace("[", "(").replace("]", ")")
    s = s.replace("{", "(").replace("}", ")")
    s = re.sub(r"\s+", "", s)
    return s


def _classify_header(col: str):
    """
    Retourne (role, sign_im)
    role in {"freq", "R", "Im", None}
    sign_im utile seulement pour Im:
        +1 si la colonne contient Im(Z)
        -1 si la colonne contient -Im(Z)
    """
    raw = str(col).strip().lower().replace("−", "-").replace("–", "-")
    s = _clean_header(col)

    # fréquence
    if "freq" in s or s in {"f", "frequency", "frequency(hz)"}:
        return "freq", None

    # partie réelle de Z
    if (
        "re(z" in raw
        or "real(z" in raw
        or "rez" in s
        or "zreal" in s
        or s == "z'"
        or s.startswith("z'")
    ):
        return "R", None

    # partie imaginaire -Im(Z)
    if (
        "-im(z" in raw
        or raw.startswith("-im")
        or "-z''" in raw
        or "-zimag" in s
    ):
        return "Im", -1

    # partie imaginaire Im(Z)
    if (
        "im(z" in raw
        or "imag(z" in raw
        or "imz" in s
        or "zimag" in s
        or "z''" in raw
    ):
        return "Im", +1

    return None, None


def _detect_header_and_sep(path: Path, max_lines: int = 120):
    """
    Cherche une ligne d'en-tête contenant freq + R + Im.
    Retourne (header_line_idx, sep)
    sep ∈ {'\\t', ';', r'\\s+'}
    """
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    for i, line in enumerate(lines[:max_lines]):
        stripped = line.strip()
        if not stripped:
            continue

        if "\t" in line:
            parts = [p.strip() for p in line.rstrip("\n\r").split("\t")]
            sep = "\t"
        elif ";" in line:
            parts = [p.strip() for p in line.rstrip("\n\r").split(";")]
            sep = ";"
        else:
            parts = re.split(r"\s+", stripped)
            sep = r"\s+"

        found = set()
        for p in parts:
            role, _ = _classify_header(p)
            if role:
                found.add(role)

        if {"freq", "R", "Im"}.issubset(found):
            return i, sep

    return None, None


def read_freq_R_Im(path: str | Path) -> pd.DataFrame:
    """
    Retourne un DataFrame normalisé avec colonnes:
        freq, R, Im

    Convention interne:
        Im = vraie partie imaginaire de Z
    Donc si le fichier source contient '-Im(Z)', on reconstruit:
        Im = -(-Im)
    """
    path = Path(path)

    header_idx, sep = _detect_header_and_sep(path)

    if header_idx is not None:
        df = pd.read_csv(
            path,
            sep=sep,
            skiprows=header_idx,
            header=0,
            dtype=str,
            engine="python",
            encoding="utf-8",
        )

        selected = {}
        im_sign = +1

        for col in df.columns:
            role, sign = _classify_header(col)
            if role and role not in selected:
                selected[role] = col
                if role == "Im":
                    im_sign = sign if sign is not None else +1

        missing = [k for k in ["freq", "R", "Im"] if k not in selected]
        if missing:
            raise ValueError(f"Colonnes introuvables: {missing}")

        out = df[[selected["freq"], selected["R"], selected["Im"]]].copy()
        out.columns = ["freq", "R", "Im"]

        for c in ["freq", "R", "Im"]:
            out[c] = out[c].astype(str).str.replace(",", ".", regex=False)
            out[c] = pd.to_numeric(out[c], errors="coerce")

        # si la colonne source était -Im, on revient à Im physique
        if im_sign == -1:
            out["Im"] = -out["Im"]

        out = out.dropna(subset=["freq", "R", "Im"]).reset_index(drop=True)
        return out

    # fallback: fichier simple 3 colonnes
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        first = f.readline().strip().split()

    def is_num(x: str) -> bool:
        try:
            float(x.replace(",", "."))
            return True
        except Exception:
            return False

    has_header = not (len(first) >= 3 and all(is_num(x) for x in first[:3]))

    df = pd.read_csv(
        path,
        sep=r"\s+",
        header=0 if has_header else None,
        names=["freq", "R", "Im"] if not has_header else None,
        dtype=str,
        engine="python",
    )

    df = df.iloc[:, :3].copy()
    df.columns = ["freq", "R", "Im"]

    for c in ["freq", "R", "Im"]:
        df[c] = df[c].astype(str).str.replace(",", ".", regex=False)
        df[c] = pd.to_numeric(df[c], errors="coerce")

    df = df.dropna(subset=["freq", "R", "Im"]).reset_index(drop=True)
    return df


def export_freq_R_Im(df: pd.DataFrame, out_path: str | Path, sep: Optional[str] = None):
    out_path = Path(out_path)

    if sep is None:
        sep = "\t" if out_path.suffix.lower() in {".tsv", ".txt"} else ","

    out = df[["freq", "R", "Im"]].copy()
    out.columns = ["freq_Hz", "Zprime_Ohm", "Zdoubleprime_Ohm"]
    out.to_csv(out_path, index=False, sep=sep)


def read_eis_raw_table(path: str | Path) -> tuple[pd.DataFrame, dict]:
    """
    Lit la table brute complète d'un fichier EIS avec ses headers d'origine.
    Retourne:
        raw_df, meta
    meta contient au minimum:
        freq_col, R_col, Im_col, im_sign
    """
    path = Path(path)

    header_idx, sep = _detect_header_and_sep(path)
    if header_idx is None:
        raise ValueError(
            "Impossible de trouver une vraie ligne d'en-tête. "
            "Le batch avec conservation des headers d'origine nécessite un fichier tabulaire avec headers."
        )

    try:
        raw_df = pd.read_csv(
            path,
            sep=sep,
            skiprows=header_idx,
            header=0,
            dtype=str,
            engine="python",
            encoding="utf-8",
        )
    except UnicodeDecodeError:
        raw_df = pd.read_csv(
            path,
            sep=sep,
            skiprows=header_idx,
            header=0,
            dtype=str,
            engine="python",
            encoding="latin1",
        )

    # On garde les noms de colonnes d'origine pour l'export.
    # Pour la détection, on travaille juste sur une vue "strippée".
    selected = {}
    im_sign = +1

    for col in raw_df.columns:
        role, sign = _classify_header(str(col).strip())
        if role and role not in selected:
            selected[role] = col
            if role == "Im":
                im_sign = sign if sign is not None else +1

    missing = [k for k in ["freq", "R", "Im"] if k not in selected]
    if missing:
        raise ValueError(f"Colonnes introuvables dans la table brute: {missing}")

    meta = {
        "freq_col": selected["freq"],
        "R_col": selected["R"],
        "Im_col": selected["Im"],
        "im_sign": im_sign,
    }
    return raw_df, meta


def split_raw_eis_by_start_freq(
    raw_df: pd.DataFrame,
    freq_col: str,
    start_freq_hz: float = 1e6,
    rel_tol: float = 0.15,
    jump_ratio: float = 5.0,
    min_points: int = 10,
) -> list[pd.DataFrame]:
    """
    Découpe la table brute en plusieurs mesures à partir de la colonne fréquence.
    Les colonnes et leurs noms d'origine sont conservés.
    """
    if freq_col not in raw_df.columns:
        raise ValueError(f"Colonne fréquence introuvable: {freq_col}")

    freq_num = pd.to_numeric(
        raw_df[freq_col].astype(str).str.replace(",", ".", regex=False),
        errors="coerce"
    )

    valid_mask = freq_num.notna()
    if valid_mask.sum() == 0:
        raise ValueError("Aucune fréquence numérique exploitable pour le découpage.")

    valid_positions = np.flatnonzero(valid_mask.to_numpy())
    freq_valid = freq_num.iloc[valid_positions].to_numpy(dtype=float)

    low = start_freq_hz * (1.0 - rel_tol)
    high = start_freq_hz * (1.0 + rel_tol)

    start_idx_valid = [0]
    for i in range(1, len(freq_valid)):
        current = freq_valid[i]
        prev = freq_valid[i - 1]

        near_start = (low <= current <= high)
        upward_jump = (prev > 0) and (current / prev >= jump_ratio)

        if near_start and upward_jump:
            start_idx_valid.append(i)

    # Convertir les indices "valides" en indices de lignes brutes
    start_rows = [valid_positions[i] for i in start_idx_valid]

    batches = []
    for k, start_row in enumerate(start_rows):
        end_row = start_rows[k + 1] if (k + 1) < len(start_rows) else len(raw_df)
        block = raw_df.iloc[start_row:end_row].copy().reset_index(drop=True)

        # Compter combien de lignes du bloc ont une fréquence numérique
        block_freq = pd.to_numeric(
            block[freq_col].astype(str).str.replace(",", ".", regex=False),
            errors="coerce"
        )
        if block_freq.notna().sum() >= min_points:
            batches.append(block)

    return batches


def build_measure_sheet_name(source_path: str | Path, measure_idx: int) -> str:
    """
    Construit un nom de feuille compact à partir du nom de fichier source.

    Exemple:
        AB_ESP1_12_03_2026_Csur100_tx0min_03_PEIS_C01.txt
    ->  ESP1_C100_tx0_03_PEIS_M1
    """
    stem = Path(source_path).stem  # enlève seulement la dernière extension

    # split sur underscore
    tokens = [t.strip() for t in stem.split("_") if t.strip()]

    cleaned = []
    i = 0
    while i < len(tokens):
        tok = tokens[i]

        # enlever préfixe AB / AB044 / etc.
        if re.fullmatch(r"AB\d*", tok, flags=re.IGNORECASE):
            i += 1
            continue

        # enlever triplet de date DD_MM_YYYY
        if (
            i + 2 < len(tokens)
            and re.fullmatch(r"\d{2}", tokens[i])
            and re.fullmatch(r"\d{2}", tokens[i + 1])
            and re.fullmatch(r"\d{4}", tokens[i + 2])
        ):
            i += 3
            continue

        # transformer Csur100 -> C100
        m = re.fullmatch(r"Csur(\d+)", tok, flags=re.IGNORECASE)
        if m:
            cleaned.append(f"C{m.group(1)}")
            i += 1
            continue

        # transformer tx0min -> tx0
        m = re.fullmatch(r"(tx[^_]*)min", tok, flags=re.IGNORECASE)
        if m:
            cleaned.append(m.group(1))
            i += 1
            continue

        # enlever channel final du type C01 / C02 / C10
        if i == len(tokens) - 1 and re.fullmatch(r"C\d{1,3}", tok, flags=re.IGNORECASE):
            i += 1
            continue

        # enlever suffixe technique .mpt resté dans stem de fichiers .mpt.txt
        if tok.lower() == "mpt":
            i += 1
            continue

        cleaned.append(tok)
        i += 1

    # option:nomage
    useful = []
    for tok in cleaned:
        if re.fullmatch(r"ESP\d+", tok, flags=re.IGNORECASE):
            useful.append(tok)
        elif re.fullmatch(r"C\d+", tok, flags=re.IGNORECASE):
            useful.append(tok)
        elif re.fullmatch(r"tx.*", tok, flags=re.IGNORECASE):
            useful.append(tok)
        elif re.fullmatch(r"\d{2}", tok):
            useful.append(tok)
        elif re.fullmatch(r"[A-Z]*EIS", tok, flags=re.IGNORECASE):
            useful.append(tok)

    # fallback si le filtrage est trop agressif
    if not useful:
        useful = cleaned if cleaned else [stem]

    base = "_".join(useful)
    sheet_name = f"{base}_M{measure_idx}"
    return sanitize_excel_sheet_name(sheet_name, fallback=f"M{measure_idx}")


def export_raw_eis_batches_to_excel(
    batches: list[pd.DataFrame],
    out_path: str | Path,
    source_path: str | Path,
    freq_col: str | None = None,
):
    """
    Exporte les mesures vers Excel en gardant:
      - toutes les colonnes
      - les headers d'origine
      - des noms de feuilles compacts
    """
    out_path = Path(out_path)

    if not batches:
        raise ValueError("Aucune mesure à exporter.")

    summary_rows = []
    for i, block in enumerate(batches, start=1):
        row = {
            "measure": i,
            "sheet_name": build_measure_sheet_name(source_path, i),
            "n_rows": len(block),
        }

        if freq_col is not None and freq_col in block.columns:
            f = pd.to_numeric(
                block[freq_col].astype(str).str.replace(",", ".", regex=False),
                errors="coerce"
            )
            f_valid = f.dropna()
            if not f_valid.empty:
                row["f_start_Hz"] = float(f_valid.iloc[0])
                row["f_end_Hz"] = float(f_valid.iloc[-1])
                row["n_freq_points"] = int(f_valid.shape[0])

        summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        used_names = set()
        for i, block in enumerate(batches, start=1):
            sheet_name = build_measure_sheet_name(source_path, i)

            # sécurité unicité
            base_name = sheet_name
            n = 2
            while sheet_name in used_names:
                suffix = f"_{n}"
                sheet_name = sanitize_excel_sheet_name(base_name[:31-len(suffix)] + suffix)
                n += 1

            used_names.add(sheet_name)
            block.to_excel(writer, sheet_name=sheet_name, index=False)


def sanitize_excel_sheet_name(name: str, fallback: str = "Sheet1") -> str:
    """
    Nettoie un nom de feuille Excel :
    - supprime les caractères interdits
    - limite à 31 caractères
    """
    name = str(name).strip()
    name = re.sub(r'[:\\/*?\[\]]', "_", name)
    name = name.strip(" '")

    if not name:
        name = fallback

    return name[:31]


def make_unique_sheet_name(candidate: str, used_names: set[str], file_stem: str | None = None) -> str:
    """
    Rend un nom de sheet unique dans le classeur cible.
    Essaie d'abord le nom d'origine, puis éventuellement un nom préfixé,
    puis ajoute un suffixe _2, _3, ...
    """
    candidate = sanitize_excel_sheet_name(candidate)

    if candidate not in used_names:
        return candidate

    if file_stem:
        prefixed = sanitize_excel_sheet_name(f"{file_stem}_{candidate}")
        if prefixed not in used_names:
            return prefixed

    base = candidate
    n = 2
    while True:
        suffix = f"_{n}"
        trial = sanitize_excel_sheet_name(base[: 31 - len(suffix)] + suffix)
        if trial not in used_names:
            return trial
        n += 1


def copy_worksheet_content(src_ws, dst_ws, copy_styles: bool = True):
    """
    Copie le contenu d'une feuille openpyxl vers une autre.
    Cette version copie :
      - valeurs / formules
      - styles principaux
      - commentaires / hyperliens
      - largeurs de colonnes / hauteurs de lignes
      - cellules fusionnées
      - freeze panes
      - autofilter
    """
    # Copier cellules
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            if copy_styles and cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)

            if cell.hyperlink:
                new_cell._hyperlink = copy(cell.hyperlink)

            if cell.comment:
                new_cell.comment = copy(cell.comment)

    # Colonnes
    for key, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[key]
        dst_dim.width = dim.width
        dst_dim.hidden = dim.hidden
        dst_dim.bestFit = dim.bestFit
        dst_dim.outlineLevel = dim.outlineLevel
        dst_dim.collapsed = dim.collapsed

    # Lignes
    for key, dim in src_ws.row_dimensions.items():
        dst_dim = dst_ws.row_dimensions[key]
        dst_dim.height = dim.height
        dst_dim.hidden = dim.hidden
        dst_dim.outlineLevel = dim.outlineLevel
        dst_dim.collapsed = dim.collapsed

    # Cellules fusionnées
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # Options diverses
    dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter and src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref

    try:
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    except Exception:
        pass


def collect_excel_workbooks(
    file_paths: list[str | Path],
    out_path: str | Path,
    copy_styles: bool = True,
):
    """
    Rassemble toutes les feuilles de plusieurs fichiers Excel dans un seul classeur.
    On ne fusionne pas les données : chaque sheet est conservée séparément.
    """
    if not file_paths:
        raise ValueError("Aucun fichier Excel fourni.")

    out_path = Path(out_path)

    dst_wb = Workbook()
    # supprimer la feuille vide par défaut
    default_ws = dst_wb.active
    dst_wb.remove(default_ws)

    used_names: set[str] = set()

    for file_path in file_paths:
        file_path = Path(file_path)
        file_stem = file_path.stem

        src_wb = load_workbook(file_path, data_only=False)

        for src_ws in src_wb.worksheets:
            target_name = make_unique_sheet_name(
                candidate=src_ws.title,
                used_names=used_names,
                file_stem=file_stem,
            )
            used_names.add(target_name)

            dst_ws = dst_wb.create_sheet(title=target_name)
            copy_worksheet_content(src_ws, dst_ws, copy_styles=copy_styles)

    dst_wb.save(out_path)


# =========================
# Warburg Wo
# =========================

def coth(z: np.ndarray) -> np.ndarray:
    return 1.0 / np.tanh(z)


def z_wo(freq_hz: np.ndarray, wor: float, wot: float, wop: float) -> np.ndarray:
    omega = 2.0 * np.pi * freq_hz.astype(float)
    arg = (1j * wot * omega) ** wop
    return wor * coth(arg) / arg


def remove_warburg_wo(
    df: pd.DataFrame,
    wor: float,
    wot: float,
    wop: float,
    fmin: Optional[float] = None,
    trim_pos_im: bool = False,
) -> pd.DataFrame:
    """
    Applique la suppression du terme Wo sur un DataFrame normalisé:
        colonnes attendues: freq, R, Im

    Retourne un DataFrame avec:
        freq, R, Im, R_meas, Im_meas, Wo_R, Wo_Im
    """
    required = {"freq", "R", "Im"}
    if not required.issubset(df.columns):
        raise ValueError("Le DataFrame doit contenir: freq, R, Im")

    freq = df["freq"].to_numpy(dtype=float)
    re_meas = df["R"].to_numpy(dtype=float)
    im_meas = df["Im"].to_numpy(dtype=float)

    z_meas = re_meas + 1j * im_meas
    zwo = z_wo(freq, wor, wot, wop)
    z_corr = z_meas - zwo

    out = pd.DataFrame({
        "freq": freq,
        "R": z_corr.real,
        "Im": z_corr.imag,
        "R_meas": re_meas,
        "Im_meas": im_meas,
        "Wo_R": zwo.real,
        "Wo_Im": zwo.imag,
    })

    mask = np.ones(len(out), dtype=bool)

    if fmin is not None:
        mask &= (out["freq"].to_numpy() >= float(fmin))

    if trim_pos_im:
        mask &= (out["Im"].to_numpy() <= 0)

    out = out.loc[mask].reset_index(drop=True)
    return out

